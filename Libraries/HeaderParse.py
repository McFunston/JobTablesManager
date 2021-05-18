from openpyxl import load_workbook
from xlrd import open_workbook
from openpyxl import Workbook
import datetime
import csv
import sys

# Splits Excel files (xls, xlsx, or csv) into multiple csv files with a single row of data (row 2) and column headers (row 1).
# requires 3 arguments:
# 1 The full path name of the file to be split (ie c:/folder/file.xls)
# 2 The ID column - The column to use as names for the files being produced - The name will be [cell.value].csv
# 3 The folder to output the csv files to


def OpenXLSX(inputPath):
    book = load_workbook(inputPath)
    return book.active


def OpenXLS(inputPath):
    with open_workbook(inputPath) as book:
        sheet = book.sheet_by_index(0)
    return sheet.get_rows()


def OpenCSV(inputPath):
    rows = []
    with open(inputPath, encoding="utf-8-sig") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            rows.append(row)
    return rows


def GetRows(sheet):
    table = list()
    for row in sheet:
        rowList = []
        for column in row:
            if type(column) != str:
                rowList.append(column.value)
            else:
                rowList.append(column)
        table.append(rowList)
    return table


def WriteXlsx(table, outputPath):
    wb = Workbook()
    ws_write = wb.active
    for row in table:
        ws_write.append(row)
    ws_write.column_dimensions['A'].width = 10
    ws_write.column_dimensions['B'].width = 40
    ws_write.column_dimensions['E'].width = 15
    ws_write.column_dimensions['F'].width = 20
    wb.save(outputPath)


def OpenTable(path):
    extension = path.split(".")[-1].lower()
    if extension == "xls":
        return OpenXLS(path)
    if extension == "xlsx":
        return OpenXLSX(path)
    if extension == "csv":
        return OpenCSV(path)


def NameParse(name):
    name = name.replace('-', '_')
    name = name.replace(' ', '_')
    splitName = name.split('_')
    for fragment in splitName:
        if fragment.isnumeric() and len(fragment) > 1:
            return fragment


def OpenHeaders(headerfile):
    with open(headerfile, 'r') as hfile:
        headerList = hfile.read().split('\n')
        newList = []
        newDim = []
        for headerItem in headerList:
            newDim.append(headerItem)
            if len(newDim) == 3:
                newList.append(newDim)
                newDim = list()

        return newList


def padJob(name):
    jn = name[1:]
    jn = jn.zfill(4)
    return name[0] + jn


def headerParse(oldFile, headerFile, idColumn, dateColumn):
    headers = OpenHeaders(headerFile)
    oldSheet = OpenTable(oldFile)
    oldTable = GetRows(oldSheet)
    found = False
    columnNames = oldTable.pop(0)
    for line in headers:
        #print(line[0])
        for oldRow in oldTable:
                # if found == True:
                #     newDate = line.split()[0]
                #     oldRow[dateColumn]=newDate
                #     found=False
            if len(oldRow[idColumn]) > 5:
                parsedName = str(NameParse(oldRow[idColumn]))
                parsedName = parsedName.zfill(4)
            else:
                parsedName = padJob(str(oldRow[idColumn]))
            if (
                parsedName != None
                and parsedName in line[0]
                and oldRow[dateColumn] is None
            ):
                print(parsedName)
                if datetime.datetime.strptime(oldRow[9], '%m/%d/%Y') <= datetime.datetime.strptime(line[1].split()[0], '%m/%d/%Y'):
                    oldRow[dateColumn] = line[1].split()[0]
                    found = True
                    break

    oldTable.insert(0, columnNames)
    WriteXlsx(oldTable, oldFile)


#headerParse('BVM_Jobs.xlsx', 'Approved.txt', 0, 3)

if __name__ == "__main__":
    if len(sys.argv) != 5:
        print(len(sys.argv))
        print(sys.argv[1])
        print("Argument: 'Excel File', 'Headers File', 'ID Column', 'DateColumn' is required")
    else:
        headerParse(sys.argv[1], sys.argv[2], int(sys.argv[3]), int(sys.argv[4]))
