from typing import Dict, List
from openpyxl import Workbook, load_workbook
from xlrd import open_workbook
from pandas import read_excel, read_csv, core
import pandas as pd

def OpenXLSX(inputPath, tab):
    book = load_workbook(inputPath)
    return book[tab] if len(book.sheetnames) > 1 else book.active


def WriteXlsx(table, outputPath):
    wb = Workbook()
    ws_write = wb.active
    headersKeep = table.pop(0)
    table.sort()
    table.insert(0, headersKeep)
    for row in table:
        ws_write.append(row)
    wb.save(outputPath)

def OpenXLS(inputPath):
    with open_workbook(inputPath) as book:
        sheet = book.sheet_by_index(0)
    return sheet.get_rows()



def GetRows(sheet):
    table = list()
    for row in sheet:
        rowList = []
        for column in row:
            if type(column) != str:
                rowList.append(column.value)
            else:
                rowList.append(column)
        table.append(rowList)
    return table

def getDict(sheet):
    entries = []
    columnHeaders = [column.value for column in sheet[1]]
    sheet.delete_rows(1, 1)
    for row in sheet:
        entry = {columnHeaders[index]: cel.value for index, cel in enumerate(row)}
        entries.append(entry)
    return entries

def GetData(path: str, tab: str) -> list:
    df = read_excel(path, tab) if tab != '' else read_excel(path)
    df1 = df.where(pd.notnull(df), None)
    dfDict = df1.to_dict('records')
    for d in dfDict:
        for c in d:       
            if str(type(d[c]))=="<class 'pandas._libs.tslibs.timestamps.Timestamp'>":
                d[c]=d[c].to_pydatetime()
            if str(type(d[c]))=="<class 'pandas._libs.tslibs.nattype.NaTType'>":
                d[c] = None
            if 'phone' not in str(c).lower():
                try:
                    d[c]=int(d[c])
                except:
                    pass

    return dfDict

def get_data_csv(path, columns) -> List:
    df: pd.DataFrame = read_csv(path, names= columns)
    df1 = df.where(pd.notnull(df), None)
    dfDict = df1.to_dict('records')
    for d in dfDict:
        for c in d:
            try:
                d[c]=int(d[c])
            except:
                pass
    return dfDict

def WriteData(path: str, sheet: str, data_list: pd.DataFrame, columns: List):
    df = pd.DataFrame(data_list)
    df.to_excel(path, sheet, columns=columns, index=False)

# test = GetData("Test_Data/BVM_Jobs.xlsx", "Sheet")
# print("stop")